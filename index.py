import os
import re
import asyncio
import aiohttp
import openpyxl
from openpyxl import Workbook
from urllib.parse import urljoin
from pathlib import Path
from PIL import Image
from io import BytesIO
from playwright.async_api import async_playwright

# Config
INPUT_EXCEL = "italy_brands.xlsx"
OUTPUT_EXCEL = "brand_banners.xlsx"
IMAGE_DIR = Path("images")
MIN_WIDTH = 600
PREFERRED_WIDTH = 720
NAV_TIMEOUT = 45000
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
)
CONCURRENCY = 5

IMAGE_DIR.mkdir(exist_ok=True)


# --- Helpers ---
def read_brands():
    wb = openpyxl.load_workbook(INPUT_EXCEL)
    sheet = wb.active
    brands = []
    for row in sheet.iter_rows(values_only=True):
        if row and row[0]:
            brands.append(str(row[0]).strip())
    return brands


def guess_homepages(brand: str):
    clean = re.sub(r"[^a-zA-Z]", "", brand).lower()
    return [
        f"https://{clean}.com",
        f"https://www.{clean}.com",
        f"https://{clean}.it",
        f"https://www.{clean}.it",
    ]


async def download_and_validate_image(session, url, brand):
    try:
        async with session.get(url, headers={"User-Agent": USER_AGENT}, timeout=30) as resp:
            if resp.status != 200:
                return None
            data = await resp.read()

        img = Image.open(BytesIO(data))
        width, height = img.size
        aspect_ratio = width / height

        if width < MIN_WIDTH or aspect_ratio < 1.0 or len(data) < 10000:
            return None

        ext = img.format.lower()
        if ext not in ["jpeg", "jpg", "png", "webp"]:
            return None

        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "", brand.replace(" ", "_"))
        filename = IMAGE_DIR / f"{safe_name}_banner.{ 'jpg' if ext=='jpeg' else ext}"
        with open(filename, "wb") as f:
            f.write(data)

        print(f"📸 {brand}: Validated {width}x{height} ({aspect_ratio:.2f})")
        return {"filename": str(filename), "width": width, "height": height}
    except Exception as e:
        print(f"❌ {brand}: Failed to download {url} → {e}")
        return None


async def extract_candidate_images(page, base_url):
    candidates = []

    await page.wait_for_timeout(3000)

    og = await page.get_attribute('meta[property="og:image"]', "content")
    if og:
        candidates.append({"url": urljoin(base_url, og), "score": 100})

    twitter = await page.get_attribute('meta[name="twitter:image"]', "content")
    if twitter:
        candidates.append({"url": urljoin(base_url, twitter), "score": 90})

    imgs = await page.query_selector_all("img")
    for img in imgs:
        src = await img.get_attribute("src")
        if not src:
            continue
        src = urljoin(base_url, src)
        box = await img.bounding_box() or {}
        width, height = box.get("width", 0), box.get("height", 0)
        score = 50
        if width > 800:
            score += 20
        if width > 1200:
            score += 10
        if width > height * 1.5:
            score += 20
        candidates.append({"url": src, "score": score})

    seen = {}
    for c in candidates:
        seen[c["url"]] = c
    return sorted(seen.values(), key=lambda x: x["score"], reverse=True)


async def try_visit_and_extract(page, url, brand, session):
    try:
        await page.goto(url, wait_until="networkidle", timeout=NAV_TIMEOUT)
    except:
        return None

    candidates = await extract_candidate_images(page, url)
    print(f"{brand}: Found {len(candidates)} candidate images")

    for c in candidates:
        res = await download_and_validate_image(session, c["url"], brand)
        if res:
            return {"brand": brand, "sourceUrl": c["url"], "localPath": res["filename"]}
    return None


async def fetch_brand_banner(brand, browser, session):
    page = await browser.new_page(user_agent=USER_AGENT, viewport={"width": 1366, "height": 768})

    for url in guess_homepages(brand):
        result = await try_visit_and_extract(page, url, brand, session)
        if result:
            await page.close()
            return result

    await page.close()
    return {"brand": brand, "sourceUrl": None, "localPath": None}


async def worker(brands, results, browser, session):
    while brands:
        brand = brands.pop(0)
        print(f"\nProcessing: {brand}")
        try:
            res = await fetch_brand_banner(brand, browser, session)
            results.append(res)
            if res["localPath"]:
                print(f"✅ {brand}: SUCCESS")
            else:
                print(f"❌ {brand}: FAILED")
        except Exception as e:
            print(f"💥 {brand}: ERROR - {e}")
            results.append({"brand": brand, "sourceUrl": None, "localPath": None})


async def main():
    brands = read_brands()
    if not brands:
        print("No brands found in Excel.")
        return

    print(f"🚀 Starting for {len(brands)} brands")

    results = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--no-sandbox"])
        async with aiohttp.ClientSession() as session:
            tasks = [
                worker(brands, results, browser, session)
                for _ in range(CONCURRENCY)
            ]
            await asyncio.gather(*tasks)
        await browser.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Banners"
    ws.append(["Brand", "SourceUrl", "LocalPath"])
    for r in results:
        ws.append([r["brand"], r["sourceUrl"], r["localPath"]])
    wb.save(OUTPUT_EXCEL)

    print(f"\n🎉 Done! Results in {OUTPUT_EXCEL}")


if __name__ == "__main__":
    asyncio.run(main())
